from openpyxl import load_workbook
import xlrd

info_workbook = xlrd.open_workbook("information.xlsx")
info_sheet = info_workbook.sheet_by_index(0)
nrows = info_sheet.nrows
for number in range(1, nrows):
    context = info_sheet.row_values(number)
    template_workbook = load_workbook('402.xlsx')
    template_sheet = template_workbook['Sheet1']

    # 采集人信息
    template_sheet.cell(4, 1).value = "采集人：{}".format(context[0])

    # 采集时间
    # template_sheet.cell(6, 1).value = context[1]

    # 房屋室号
    address_string = context[2].split('-')
    template_sheet.cell(3, 3).value = address_string[0]
    template_sheet.cell(3, 5).value = address_string[1]
    template_sheet.cell(3, 6).value = address_string[2]
    template_sheet.cell(3, 7).value = address_string[3]

    # 姓名
    template_sheet.cell(4, 4).value = context[3]
    # 身份证号
    template_sheet.cell(4, 6).value = "身份证:  {}".format(str(context[4]).split('.')[0])
    # 护照
    template_sheet.cell(5, 6).value = "护照:  {}".format(str(context[5]).split('.')[0])
    # 证件号
    template_sheet.cell(6, 4).value = context[6]
    # 居住地
    template_sheet.cell(8, 4).value = context[7]
    # 房主电话
    template_sheet.cell(8, 8).value = str(context[8]).split('.')[0]

    # 单位名称
    template_sheet.cell(9, 4).value = context[9]
    # 治安-责任姓名
    template_sheet.cell(11, 4).value = context[10]
    # 治安-身份证号
    template_sheet.cell(11, 6).value = context[11]
    # 治安-手机号
    template_sheet.cell(12, 6).value = context[12]
    # 治安-现住地
    template_sheet.cell(13, 6).value = context[13]

    # 房屋类型
    horse_type = context[14]
    if horse_type == "自住房屋":
        template_sheet.cell(15, 2).value = "√"
    elif horse_type == "出租房屋":
        template_sheet.cell(15, 4).value = "√"
    elif horse_type == "闲置空房":
        template_sheet.cell(15, 6).value = "√"
    elif horse_type == "单位宿舍":
        template_sheet.cell(15, 8).value = "√"
    elif horse_type == "商业用房":
        template_sheet.cell(17, 2).value = "√"
    elif horse_type == "工地工棚":
        template_sheet.cell(17, 4).value = "√"
    elif horse_type == "田间窝棚":
        template_sheet.cell(17, 6).value = "√"
    else:
        template_sheet.cell(17, 8).value = "√"

    template_workbook.save('{}.xlsx'.format("".join(address_string)))
