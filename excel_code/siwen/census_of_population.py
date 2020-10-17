# -*- coding:utf-8 -*-
import os

from openpyxl import load_workbook
import xlrd

addresses = []


def set_person_info(context, sheet, number):
    start_line = 20
    index = start_line + 2*number
    sheet.cell(index, 1).value = number
    sheet.cell(index, 2).value = context[0]
    sheet.cell(index, 3).value = context[2]
    sheet.cell(index, 4).value = context[1]
    sheet.cell(index, 5).value = context[3]
    sheet.cell(index, 7).value = context[5]


def generate_excel(information):
    info_workbook = xlrd.open_workbook(information)
    info_sheet = info_workbook.sheet_by_index(0)
    nrows = info_sheet.nrows
    for number in range(1, nrows):
        context = info_sheet.row_values(number)
        template_workbook = load_workbook('402.xlsx')
        template_sheet = template_workbook['Sheet1']
        if context[5] and "盘龙小区" not in context[5] and "唐城街" not in context[5] and "四合加苑" not in context[5] and "洪南路245号" not in context[5]:
            print(context)
            address = context[5]
            if address in addresses:
                obj_file = os.path.join("tables", '{}.xlsx'.format(address))
                workbook = load_workbook(obj_file)
                sheet = workbook['Sheet1']
                no = None
                for index in range(22, 34, 2):
                    number = sheet.cell(index, 1).value
                    if number:
                        no = number
                    else:
                        break
                set_person_info(context, sheet, no+1)
                workbook.save(obj_file)
                continue
            else:
                addresses.append(address)

            template_sheet.cell(3, 3).value = address
            set_person_info(context, template_sheet, 1)
            file_pth = os.path.join("tables", '{}.xlsx'.format(address))
            template_workbook.save(file_pth)



        # # 房屋室号
        # address_string = context[2].split('-')
        # template_sheet.cell(3, 3).value = address_string[0]
        # template_sheet.cell(3, 5).value = address_string[1]
        # template_sheet.cell(3, 6).value = address_string[2]
        # template_sheet.cell(3, 7).value = address_string[3]
        #
        # # 姓名
        # template_sheet.cell(4, 4).value = context[3]
        # # 身份证号
        # template_sheet.cell(4, 6).value = "身份证:  {}".format(str(context[4]).split('.')[0])
        # # 护照
        # template_sheet.cell(5, 6).value = "护照:  {}".format(str(context[5]).split('.')[0])
        # # 证件号
        # template_sheet.cell(6, 4).value = context[6]
        # # 居住地
        # template_sheet.cell(8, 4).value = context[7]
        # # 房主电话
        # template_sheet.cell(8, 8).value = str(context[8]).split('.')[0]
        #
        # # 单位名称
        # template_sheet.cell(9, 4).value = context[9]
        # # 治安-责任姓名
        # template_sheet.cell(11, 4).value = context[10]
        # # 治安-身份证号
        # template_sheet.cell(11, 6).value = context[11]
        # # 治安-手机号
        # template_sheet.cell(12, 6).value = context[12]
        # # 治安-现住地
        # template_sheet.cell(13, 6).value = context[13]
        #
        # # 房屋类型
        # horse_type = context[14]
        # if horse_type == "自住房屋":
        #     template_sheet.cell(15, 2).value = "√"
        # elif horse_type == "出租房屋":
        #     template_sheet.cell(15, 4).value = "√"
        # elif horse_type == "闲置空房":
        #     template_sheet.cell(15, 6).value = "√"
        # elif horse_type == "单位宿舍":
        #     template_sheet.cell(15, 8).value = "√"
        # elif horse_type == "商业用房":
        #     template_sheet.cell(17, 2).value = "√"
        # elif horse_type == "工地工棚":
        #     template_sheet.cell(17, 4).value = "√"
        # elif horse_type == "田间窝棚":
        #     template_sheet.cell(17, 6).value = "√"
        # else:
        #     template_sheet.cell(17, 8).value = "√"
        #
        #     set_person_info(context[15], template_sheet, 22)
        #     set_person_info(context[16], template_sheet, 24)
        #     set_person_info(context[17], template_sheet, 26)
        #     set_person_info(context[18], template_sheet, 28)
        #     set_person_info(context[19], template_sheet, 30)
        #
        # template_workbook.save('{}.xlsx'.format("".join(address_string)))


excel_files = os.listdir("20201016")
for file in excel_files:
    path = os.path.join("20201016", file)
    print(path, "=============>")
    generate_excel(path)


print("共计excel：", len(os.listdir("tables")))
