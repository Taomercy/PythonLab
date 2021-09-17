# -*- coding:utf-8 -*-
from collections import Counter

from openpyxl import load_workbook


def get_data(filename, sheet_name="all", rows=None):
    def get_data_in_sheet(sheet_name, data):
        sheet = workbook[sheet_name]
        for row in range(1, sheet.max_row + 1):
            row_data = []
            try:
                for index in rows:
                    row_data.append(sheet[row][index].value)
            except:
                print("error:", sheet_name, row)
                continue
            # TODO: filter data
            if row_data[-2] and row_data[-1] and type(row_data[-1] is int):
                data.append(row_data)
                print(row, row_data)

    if rows is None:
        rows = [0]
    assert type(rows) is list
    data = []
    workbook = load_workbook(filename)
    if sheet_name != "all":
        get_data_in_sheet(sheet_name, data)
    else:
        print(workbook.sheetnames)
        for sheet_name in workbook.sheetnames:
            get_data_in_sheet(sheet_name, data)
    return data


def write_data(fliename, sheet_name, data):
    workbook = load_workbook(fliename)
    sheet = workbook[sheet_name]
    for row in range(1, sheet.max_row+1):
        name = sheet[row][2].value
        id_card = sheet[row][3].value
        # TODO: match data and write
        # sheet.cell(x, y).value = "√"
        for i in data:
            if id_card == i[3]:
                try:
                    sheet.cell(row, 6).value = str(i[-1])
                except:
                    print("error:", row, i)
                    continue
                print(row, i)
    workbook.save("result.xlsx")


data = get_data('洪庙一居委居住人员(2021版）.xlsx', "all", [2, 3, 4, 5, 6])
print("数据总数:", len(data))
write_data("疫苗接种统计.xlsx", "Sheet1", data)




