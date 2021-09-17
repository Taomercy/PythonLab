# -*- coding:utf-8 -*-
from openpyxl import load_workbook


def get_data(filename, sheet_name, rows=None):
    if rows is None:
        rows = [0]
    assert type(rows) is list
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]
    data = []
    for row in range(1, sheet.max_row+1):
        row_data = []
        for index in rows:
            row_data.append(sheet[row][index].value)
        data.append(row_data)
        print(row_data)
    return data


def write_data(fliename, sheet_name, data):
    workbook = load_workbook(fliename)
    sheet = workbook[sheet_name]
    for row in range(1, sheet.max_row+1):
        name = sheet[row][2].value
        for i in data:
            if name == i[0]:
                try:
                    sheet.cell(row, 8).value = "√"
                except:
                    print("error:", row, i)
                    continue
                print(row, i)
        # TODO: match data and write
        # sheet.cell(x, y).value = "√"
    workbook.save("result.xlsx")


data = get_data('表三.xlsx', "已接种第一针", [1, 10, 11])
write_data("表一.xlsx", "Sheet1", data)




