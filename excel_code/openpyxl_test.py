from openpyxl import Workbook
aa = []
with open("aaa.txt", "r") as f:
     content = f.readlines()
     for i in content:
         if not i == "\n":
            i = i.replace("\n","")
            aa.append(i)


template_workbook = Workbook()
sheet = template_workbook["Sheet"]
count = 1
line = 1
print(len(aa)/3)

for i in aa:
    if count == 3:
        sheet.cell(line, count).value = i
        line += 1
        count = 1
        continue
    sheet.cell(line, count).value = i
    count += 1

template_workbook.save("bb.xlsx")