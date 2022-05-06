from openpyxl import load_workbook

work_book_name = "failed_test_suits.xlsx"

workbook = load_workbook(work_book_name)
sheet_ts_slave = workbook['TestSuit-Slave']
try:
    sheet_slaves_count = workbook['Slaves-Count']
except:
    sheet_slaves_count = workbook.create_sheet('Slaves-Count', 4)


def get_row(name, sheet):
    for i in range(sheet.max_row):
        if name == sheet.cell(i + 2, 1).value:
            return i + 2
    return sheet.max_row


def get_column(name, sheet):
    for i in range(sheet.max_column):
        if name == sheet.cell(1, i + 2).value:
            return i + 2
    return sheet.max_column


for slave_row in range(2, sheet_ts_slave.max_row+1):
    slaves = []
    ts_name = sheet_ts_slave.cell(slave_row, 1).value
    for col in range(2, sheet_ts_slave.max_column+1):
            slave = sheet_ts_slave.cell(slave_row, col).value
            slaves.append(slave)

    count = {}
    for slave in slaves:
        if slave:
            if count.get(slave):
                count[slave] += 1
            else:
                count[slave] = 1

    print(ts_name, count)
    count_row = get_row(ts_name, sheet_slaves_count)
    sheet_slaves_count.cell(count_row, 1).value = ts_name
    for s, c in count.items():
        count_column = get_column(s, sheet_slaves_count)
        sheet_slaves_count.cell(1, count_column).value = s
        sheet_slaves_count.cell(count_row, count_column).value = c

sum_max_row = None
sum_max_column = None

for i in range(2, sheet_slaves_count.max_row+2):
    if sheet_slaves_count.cell(i, 1).value is None:
        data = sheet_slaves_count.cell(i-1, 1)
        sum_max_row = data.row
        break

for i in range(2, sheet_slaves_count.max_column+2):
    if sheet_slaves_count.cell(1, i).value is None:
        data = sheet_slaves_count.cell(1, i-1)
        sum_max_column = data.column
        break

data = sheet_slaves_count.cell(sum_max_row, sum_max_column)
max_column = data.coordinate.replace(str(data.row), "")

for i in range(2, data.row):
    formula = "=SUM(B{0}:{1}{0})".format(i, max_column)
    sheet_slaves_count.cell(i, int(data.column)+1, value=formula)

for i in range(2, data.column+1):
    d = sheet_slaves_count.cell(sum_max_row, i)
    column = d.coordinate.replace(str(d.row), "")
    formula = "=SUM({0}2:{0}{1})".format(column, sum_max_row)
    sheet_slaves_count.cell(sum_max_row+1, i, value=formula)

workbook.save(work_book_name)
