from openpyxl import Workbook
from openpyxl import load_workbook

a = [{"name": "jack", "score": 12},
     {"name": "marry", "score": 14},
     {"name": "peter", "score": 18}
]

b = [{"name": "jack", "score": 19},
     {"name": "marry", "score": 23},
     {"name": "tom", "score": 34},
     {"name": "tonny", "score": 22},
]

work_book = Workbook()
sheet = work_book.create_sheet("Sheet1", 0)
date = "2022-4-13"
for i in range(len(a)):
    sheet.cell(i+2, 1).value = a[i]["name"]
    sheet.cell(1, 2).value = date
    sheet.cell(i+2, 2).value = a[i]["score"]
    print(a[i]["name"], i+2)
work_book.save("123.xlsx")


date = "2022-4-14"
work_book = load_workbook("123.xlsx")
sheet = work_book["Sheet1"]

def get_row(name, sheet):
    for i in range(sheet.max_row):
        if name == sheet.cell(i+2, 1).value:
            return i+2
    return sheet.max_row

column = sheet.max_column
sheet.cell(1, column+1).value = date
for i in range(len(b)):
    row = get_row(b[i]["name"], sheet)
    sheet.cell(row, 1).value = b[i]["name"]
    sheet.cell(row, column+1).value = b[i]["score"]
    print(b[i]["name"], row)

work_book.save("123.xlsx")