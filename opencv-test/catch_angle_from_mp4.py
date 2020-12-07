import os
import cv2
import math
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
template_workbook = Workbook()
sheet = template_workbook["Sheet"]
align = Alignment(horizontal='left', vertical='center')

global INDEX
INDEX = 0
#统计概率霍夫线变换
def line_detect_possible_demo(image):
    global INDEX
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    edges = cv2.Canny(gray, 300, 550, apertureSize=3)  # apertureSize参数默认其实就是3
    edges_image = 'edges%d.jpg' % INDEX
    if not os.path.exists("edges"):
        os.mkdir("edges")
    edges_image = os.path.join("edges", edges_image)
    cv2.imwrite(edges_image, edges)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, 60, minLineLength=60, maxLineGap=5)
    result_lines = []
    points = []
    for line in lines:
        x1, y1, x2, y2 = line[0]
        if abs(y1 - y2) < 5:
            continue
        result_lines.append(line)
        # print("coordinate:", x1, y1, x2, y2)
        points.append([x1, y1, x2, y2])
        cv2.line(image, (x1, y1), (x2, y2), (0, 0, 255), 2)
    fileName = 'detect%d.jpg' % INDEX
    if not os.path.exists("detect"):
        os.mkdir("detect")
    fileName = os.path.join("detect", fileName)
    sheet.cell(INDEX, 1).value = 'detect%d.jpg' % INDEX
    sheet.cell(INDEX, 1).alignment = align
    cv2.imwrite(fileName, image)
    return points


def angle(v1, v2):
    dx1 = v1[2] - v1[0]
    dy1 = v1[3] - v1[1]
    dx2 = v2[2] - v2[0]
    dy2 = v2[3] - v2[1]
    angle1 = math.atan2(dy1, dx1)
    angle1 = int(angle1 * 180 / math.pi)
    # print(angle1)
    angle2 = math.atan2(dy2, dx2)
    angle2 = int(angle2 * 180 / math.pi)
    # print(angle2)
    if angle1 * angle2 >= 0:
        included_angle = abs(angle1 - angle2)
    else:
        included_angle = abs(angle1) + abs(angle2)
        if included_angle > 180:
            included_angle = 360 - included_angle
    return included_angle


def picture_cal(image_path):
    src = cv2.imread(image_path)  # 调用上一个函数后，会把传入的src数组改变，所以调用下一个函数时，要重新读取图片
    src = src[120:850, 400:500]
    points = line_detect_possible_demo(src)
    global INDEX
    sheet.cell(INDEX, 3).value = str(points)
    sheet.cell(INDEX, 3).alignment = align
    try:
        first_line = points[0]
        second_line = points[1]
        ang1 = angle(first_line, second_line)
        sheet.cell(INDEX, 2).value = ang1
        sheet.cell(INDEX, 2).alignment = align
        return ang1
    except:
        sheet.cell(INDEX, 2).value = "None"
        sheet.cell(INDEX, 2).alignment = align
        return points



def main(mp4):
    cap = cv2.VideoCapture(mp4)  # 获取一个视频打开cap
    isOpened = cap.isOpened  # 判断是否打开
    # print(isOpened)
    fps = cap.get(cv2.CAP_PROP_FPS)
    # print(fps)
    # 获取宽度
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    # 获取高度
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    i = 0
    while (isOpened):
        i += 1
        (flag, frame) = cap.read()  # 读取每一帧，一张图像flag 表明是否读取成果 frame内容
        # flag表示是否成功读图
        if flag is True:
            # 控制质量
            if i % 3 == 0:
                global INDEX
                INDEX += 1
                fileName = 'catch' + str(INDEX) + '.jpg'
                if not os.path.exists("catch"):
                    os.mkdir("catch")
                fileName = os.path.join("catch", fileName)
                cv2.imwrite(fileName, frame)
                try:
                    angle = picture_cal(fileName)
                    if type(angle) is int:
                        print("\r夹角:", angle, end="", flush=True)
                except:
                    sheet.cell(INDEX, 1).value = 'detect%d.jpg' % INDEX
                    sheet.cell(INDEX, 1).alignment = align
                    sheet.cell(INDEX, 2).value = "None"
                    sheet.cell(INDEX, 2).alignment = align
                    sheet.cell(INDEX, 3).value = "未检测到直线"
                    sheet.cell(INDEX, 3).alignment = align
                    continue
        else:
            break
    cap.release()
    print('\nend!')


if __name__ == '__main__':
    try:
        os.remove("statistic.xlsx")
    except:
        pass

    main('123.mp4')
    template_workbook.save("statistic.xlsx")




