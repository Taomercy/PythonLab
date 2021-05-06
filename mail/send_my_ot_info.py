#!/usr/bin/env python
# -*- coding:utf-8 -*-\
from openpyxl import load_workbook
import datetime
import time
import sys
from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.header import Header
import getpass

br_name = "Wei"
MONTH = datetime.datetime.now().month
YEAR = datetime.datetime.now().year
mdict = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
         5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
         9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


try:
    mail_password = getpass.getpass("Please input your email password:")
except Exception as e:
    print(e)
    sys.exit()


class MailTable(object):
    html = None
    table_head = None
    td_head = None

    def __init__(self):
        self.table_head = "<table class=MsoNormalTable border=0 cellspacing=0 cellpadding=0 width=756 style='width:567.0pt;margin-left:-.65pt;border-collapse:collapse'>"
        self.td_head = """
                    <td width=64 nowrap valign=bottom style='width:48.0pt;border:solid windowtext 1.0pt;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>HRID<o:p></o:p></span></b></p></td>
                    <td width=96 valign=bottom style='width:1.0in;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>Spell Name<o:p></o:p></span></b></p></td>
                    <td width=119 nowrap valign=bottom style='width:89.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>Chinese Name<o:p></o:p></span></b></p></td>
                    <td width=133 nowrap valign=bottom style='width:100.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>PCODE<o:p></o:p></span></b></p></td>
                    <td width=97 nowrap valign=bottom style='width:73.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>OT Month<o:p></o:p></span></b></p></td>
                    <td width=81 nowrap valign=bottom style='width:61.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>OT Day<o:p></o:p></span></b></p></td>
                    <td width=80 nowrap valign=bottom style='width:60.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>OT hours<o:p></o:p></span></b></p></td>
                    <td width=85 nowrap valign=bottom style='width:64.0pt;border:solid windowtext 1.0pt;border-left:none;background:#17375D;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><b><span style='font-size:10.0pt;font-family:"Arial",sans-serif;color:white'>Event<o:p></o:p></span></b></p></td>
        """
        self.html = "<p>Hi,</p><p>My OT info in {0}：</p>".format(mdict[MONTH])
        self.html += self.table_head
        self.html += self.td_head

    def add_tr(self, spell_name, chinese_name, pcode, ot_month, ot_day, ot_hours, event):
        td = """
                    <td width=64 nowrap style='width:48.0pt;border:solid windowtext 1.0pt;border-top:none;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='color:black'>60712<o:p></o:p></span></p></td>
                    <td width=96 nowrap style='width:1.0in;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>{0}<o:p></o:p></span></p></td>
                    <td width=119 nowrap style='width:89.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span lang=ZH-CN style='font-size:10.0pt;font-family:SimSun'>{1}</span><span style='font-size:10.0pt;font-family:SimSun'><o:p></o:p></span></p></td>
                    <td width=133 nowrap valign=bottom style='width:100.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='color:black'>{2}</span><span style='font-size:10.0pt;font-family:"Times New Roman",serif'><o:p></o:p></span></p></td>
                    <td width=97 nowrap valign=bottom style='width:73.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal align=center style='text-align:center'><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>{3}<o:p></o:p></span></p></td>
                    <td width=81 nowrap valign=bottom style='width:61.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>{4}<o:p></o:p></span></p></td>
                    <td width=80 nowrap valign=bottom style='width:60.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>{5}<o:p></o:p></span></p></td>
                    <td width=85 nowrap valign=bottom style='width:64.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:#FCD5B4;padding:0in 5.4pt 0in 5.4pt;height:15.65pt'><p class=MsoNormal><span style='color:black'>{6}</span></p></td>
        """.format(spell_name, chinese_name, pcode, ot_month, ot_day, ot_hours, event)
        tr = "<tr style='height:15.0pt'>" + td + "</tr>"
        self.html += tr

    def add_total_tr(self, total):
        td = """            
            <td width=64 nowrap style='width:48.0pt;border:solid windowtext 1.0pt;border-top:none;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
            <td width=96 nowrap style='width:1.0in;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
            <td width=119 nowrap style='width:89.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:white;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
            <td width=133 nowrap valign=bottom style='width:100.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
            <td width=97 nowrap valign=bottom style='width:73.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
            <td width=81 nowrap valign=bottom style='width:61.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>Total:<o:p></o:p></span></p></td>
            <td width=80 nowrap valign=bottom style='width:60.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'><p class=MsoNormal><span style='font-size:10.0pt;font-family:"Times New Roman",serif'>{0}<o:p></o:p></span></p></td>
            <td width=85 nowrap valign=bottom style='width:64.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;background:#FCD5B4;padding:0in 5.4pt 0in 5.4pt;height:15.0pt'></td>
        """.format(total)
        tr = "<tr style='height:15.0pt'>" + td + "</tr>"
        self.html += tr

    def get_html(self):
        self.html += '</table>'
        self.html += '<p>BR/{0}</p>'.format(br_name)
        return self.html


def get_delta_time(value):
    date = value.split('  ')[0]
    timedelta = value.split('  ')[1]
    timestart = timedelta.split('-')[0]
    timeend = timedelta.split('-')[1]
    ts = date + ' ' + timestart
    te = date + ' ' + timeend
    timeArray1 = time.strptime(ts, "%m/%d/%Y %H:%M")
    timeStamp1 = int(time.mktime(timeArray1))
    dateArray1 = datetime.datetime.utcfromtimestamp(timeStamp1)

    timeArray2 = time.strptime(te, "%m/%d/%Y %H:%M")
    timeStamp2 = int(time.mktime(timeArray2))
    dateArray2 = datetime.datetime.utcfromtimestamp(timeStamp2)

    delta = dateArray2 - dateArray1
    return date, delta


def get_info():
    data = load_workbook("ot.xlsx")
    table = data['Sheet1']
    rows = table.max_row
    row_data = []
    events = []
    for i in range(2, rows + 1):
        cell_value = table.cell(row=i, column=1).value
        date = cell_value.split('  ')[0]
        if date.startswith(str(MONTH)) and date.endswith(str(YEAR)):
            row_data.append(cell_value.strip())
            events.append(table.cell(row=i, column=2).value)
    return row_data, events


def send_mail(html):
    mail_info = {
        "from": "wei.wu@cienet.com.cn",
        "to": ["wei.wu@cienet.com.cn"],
        "cc": ["1224355271@qq.com"],
        "hostname": "smtp.263xmail.com",
        "username": "wei.wu@cienet.com.cn",
        "password": mail_password,
        "mail_subject": "RE: OT hours in {0}".format(mdict[MONTH]),
        "mail_text": html,
        "mail_encoding": "utf-8"
    }

    smtp = SMTP_SSL(mail_info["hostname"])
    smtp.set_debuglevel(1)

    smtp.ehlo(mail_info["hostname"])
    smtp.login(mail_info["username"], mail_info["password"])
    msg = MIMEText(mail_info["mail_text"], _subtype="html", _charset=mail_info["mail_encoding"])
    msg["Subject"] = Header(mail_info["mail_subject"], mail_info["mail_encoding"])
    msg["From"] = mail_info["from"]
    msg["To"] = ",".join(mail_info["to"])
    msg["Cc"] = ",".join(mail_info["cc"])

    smtp.sendmail(mail_info["from"], mail_info["to"] + mail_info["cc"], msg.as_string())
    smtp.quit()


mt = MailTable()
values, events = get_info()
total_time = 0
for value, event in zip(values, events):
    date, delta = get_delta_time(value)
    print(date, delta.seconds/60.0/60.0)
    dlist = date.split('/')
    month = dlist[0]
    day = dlist[1]
    hours = delta.seconds/60.0/60.0
    total_time += hours
    mt.add_tr("Wei", "吴威", "ERIC-Shanghai-HSS", ot_month=month, ot_day=day, ot_hours=hours, event=event)
mt.add_total_tr(total_time)
#send_mail(mt.get_html())

